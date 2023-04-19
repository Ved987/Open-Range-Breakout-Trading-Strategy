from openpyxl import load_workbook
from datetime import time


def main():
    loc = "January2020.xlsx"
    workbook = load_workbook(loc)
    spreadsheet = workbook.active
    T = ()
    net_profit, net_loss = 0, 0
    Text2 = input("Enter the Stock no. (eg- 1, 2, 3, 4, All)\n")
    for v in spreadsheet.iter_cols(min_col=3, max_col=3, values_only=True):
        T = v
    if Text2 == "1":
        Stock_1 = ()
        for value in spreadsheet.iter_cols(min_col=4, max_col=4, values_only=True):
            Stock_1 = value
        net_profit, net_loss = Calculation(Stock_1, 1, T)
    elif Text2 == "2":
        Stock_2 = ()
        for value in spreadsheet.iter_cols(min_col=5, max_col=5, values_only=True):
            Stock_2 = value
        net_profit, net_loss = Calculation(Stock_2, 2, T)
    elif Text2 == "3":
        Stock_3 = ()
        for value in spreadsheet.iter_cols(min_col=6, max_col=6, values_only=True):
            Stock_3 = value
        net_profit, net_loss = Calculation(Stock_3, 3, T)
    elif Text2 == "4":
        Stock_4 = ()
        for value in spreadsheet.iter_cols(min_col=7, max_col=7, values_only=True):
            Stock_4 = value
        net_profit, net_loss = Calculation(Stock_4, 4, T)
    elif Text2 == "All":
        Stock_1 = ()
        for value in spreadsheet.iter_cols(min_col=4, max_col=4, values_only=True):
            Stock_1 = value
        profit, loss = Calculation(Stock_1, 1, T)
        net_profit = net_profit + profit
        net_loss = net_loss + loss
        Stock_2 = ()
        for value in spreadsheet.iter_cols(min_col=5, max_col=5, values_only=True):
            Stock_2 = value
        profit, loss = Calculation(Stock_2, 2, T)
        net_profit = net_profit + profit
        net_loss = net_loss + loss
        Stock_3 = ()
        for value in spreadsheet.iter_cols(min_col=6, max_col=6, values_only=True):
            Stock_3 = value
        profit, loss = Calculation(Stock_3, 3, T)
        net_profit = net_profit + profit
        net_loss = net_loss + loss
        Stock_4 = ()
        for value in spreadsheet.iter_cols(min_col=7, max_col=7, values_only=True):
            Stock_4 = value
        profit, loss = Calculation(Stock_4, 4, T)
        net_profit = net_profit + profit
        net_loss = net_loss + loss
    print("The profit earned is: \t" + str(net_profit))
    print("The loss gained is: \t" + str(net_loss))
    print("The net income earned is: \t" + str((net_profit + net_loss)))


def Open_Range(Stock, i, Time):
    T1 = time(9, 15, 0)
    T2 = time(9, 30, 0)
    Highest = Stock[0]
    Lowest = Stock[0]
    while T1 <= Time[i] <= T2:
        if Stock[i] > Highest:
            Highest = Stock[i]
        elif Stock[i] < Lowest:
            Lowest = Stock[i]
        i = i + 1
    return Highest, Lowest, i


def Buy_Sell(Stock, high, low, i, Time):
    T1 = time(9, 30, 0)
    T2 = time(15, 30, 0)
    T3 = time(15, 15, 0)
    count = 0
    bought = 0
    sold = 0
    Buy_price = 0
    Sell_price = 0
    Buy_Trade = 0
    Sell_Trade = 0
    while T1 <= Time[i] <= T2:
        # BUY TRADING
        if Stock[i] > high and bought == 0 and Sell_Trade == 0 and Buy_Trade == 0:
            i = i + 1
            count = count + 1
            if count == 10:
                bought = 1
                Buy_price = Stock[i]
                Buy_Trade = 1
                count = 0
        elif Stock[i] < (high - high * 0.005) and bought == 1 and Buy_Trade == 1:
            i = i + 1
            count = count + 1
            if count == 10:
                bought = 0
                Sell_price = Stock[i]
                count = 0
        # SELL TRADING:
        elif Stock[i] < low and sold == 0 and Buy_Trade == 0 and Sell_Trade == 0:
            i = i + 1
            count = count + 1
            if count == 10:
                sold = 1
                Sell_price = Stock[i]
                Sell_Trade = 1
                count = 0
        elif Stock[i] > (low + low * 0.005) and sold == 1 and Sell_Trade == 1:
            i = i + 1
            count = count + 1
            if count == 10:
                sold = 0
                Buy_price = Stock[i]
                count = 0
        # Square off
        elif Time[i] >= T3:
            if Buy_Trade == 1 and bought == 1:
                Sell_price = Stock[i]
                bought = 0
            elif Sell_Trade == 1 and sold == 1:
                Buy_price = Stock[i]
                sold = 0
            i = i + 1
        else:
            i = i + 1
    income = Sell_price - Buy_price
    return income, i


def Calculation(Stock, n, Time):
    print(" For Stock No.:" + str(n))
    i = 0
    profit = 0
    loss = 0
    T1 = time(9, 15, 0)
    T2 = time(9, 30, 0)
    while i in range(0, len(Time)):
        if T1 < Time[i] < T2:
            S1_high, S1_low, i = Open_Range(Stock, i, Time)
            income, i = Buy_Sell(Stock, S1_high, S1_low, i, Time)
            if income > 0:
                profit = profit + income
            elif income < 0:
                loss = loss + income
        i = i + 3
    return profit, loss


main()
